"""import os, shutil
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook

# Constants
NULL_VALUE = -99999

# BOREHOLE_INTERVAL_COLLECTION tab
COLLECTION_NAME_POSTFIX = '_NMR_results'
COLLECTION_TYPE = 'hydrogeological'
ORIGINATOR_NUMBER = 192     # Peljo, M.

# DOWNHOLE INTERVALS tab
INTERVAL_UNIT_OF_MEASURE = 'metre'

# SAMPLES tab
SAMPLE_ID_INFIX = '_NMR_'
ANO = 228       # Generic GA ANO
ACTIVITY_CODE = 'A'
SAMPLE_TYPE = 'wireline log'
SAMPLING_METHOD = 'borehole logging tool'
MATERIAL_CLASS = 'rock'

# SCALAR PROPERTIES
ACCESS_CODE = 'O'
QA_STATUS = 'U'
ORIGNO = 192
SOURCE_TYPE = 'GA NAS'
SOURCE = r'\\prod.lan\active\proj\dcd\Data\FieldworkAndPhotos\FieldData\Javelin_NMR'
LOAD_APPROVED = 'Y'
PROCESS_TYPE = ''
# INSTRUMENT_TYPE = (110, 'magnetic_resonance_system', 'magnetic_resonance_system')

# PROCESS_LOOKUP key is Javelin results table heading,
# value is (PROCESS TYPE, PETROPHYSICAL PROPERTY, UOM, RESULT QUALIFIER, UNCERTAINTY TYPE, REMARKS)
PROCESS_LOOKUP_JP350 = {
    #'depth:
    #'totalf:   ()
    'clayf':    (4439109, 'clay-bound water content', 'volume fraction', 'weighted mean', 'unknown',
                '3 ms T2 cutoff',),
    'capf':     (4439113, 'capillary-bound water content', 'volume fraction', 'weighted mean', 'unknown',
                '33 ms T2 cutoff',),
    'freef':    (4439121, 'free water content', 'volume fraction', 'weighted mean', 'unknown', '',),
    #'mlT2':    (4439125, 'mean log T2', 'second', 'weighted mean', 'unknown','',),
    'Ksdr':     (4439129, 'hydraulic conductivity', 'metres per day', 'weighted mean', 'unknown',
                'Hydraulic conductivity calculated using the Schlumberger Doll Research equation (Ksdr)',),
    #'Ktc':     (4439129, 'hydraulic conductivity', 'metres per day', 'weighted mean', 'unknown',
    #            'Hydraulic conductivity calculated using the Timur-Coates equation (Ktc)',),
    #'Ksoe':    (4439129, 'hydraulic conductivity', 'metres per day', 'weighted mean', 'unknown', '',),
    #'Tsdr':    ('', '', '', '', '', '',),
    #'Ttc':     ('', '', '', '', '', '',),
    #'Tsoe':    ('', '', '', '', '', '',),
    #'soe':     ('', '', '', '', '', '',),
    #'noise':   ('', '', '', '', '', '',),
}

PROCESS_LOOKUP_JP238 = {
    #'depth:
    #'totalf:   ()
    'clayf':    (4439106, 'clay-bound water content', 'volume fraction', 'weighted mean', 'unknown',
                '3 ms T2 cutoff',),
    'capf':     (4439111, 'capillary-bound water content', 'volume fraction', 'weighted mean', 'unknown',
                '33 ms T2 cutoff',),
    'freef':    (4439119, 'free water content', 'volume fraction', 'weighted mean', 'unknown', '',),
    #'mlT2':    (4439123, 'mean log T2', 'second', 'weighted mean', 'unknown', '',),
    'Ksdr':     (4439127, 'hydraulic conductivity', 'metres per day', 'weighted mean', 'unknown',
                'Hydraulic conductivity calculated using the Schlumberger Doll Research equation (Ksdr)',),
    #'Ktc':     (4439127, 'hydraulic conductivity', 'metres per day', 'weighted mean', 'unknown',
    #            'Hydraulic conductivity calculated using the Timur-Coates equation (Ktc)',),
    #'Ksoe':    (4439127, 'hydraulic conductivity', 'metres per day', 'weighted mean', 'unknown', '',),
    #'Tsdr':    ('', '', '', '', '', '',),
    #'Ttc':     ('', '', '', '', '', '',),
    #'Tsoe':    ('', '', '', '', '', '',),
    #'soe':     ('', '', '', '', '', '',),
    #'noise':   ('', '', '', '', '', '',),
}

# Read template and input data config
full_template_file = r'.\Templates\2023_05_24_Javelin_bores_for_loading-CompletedSurvey.xlsx'
template_file = r'.\Templates\ROCKPROPS-jav_cols.XLSX'
metadata_file = r'.\Config\2023_05_24_Javelin_bores_for_testing.xlsx'

# Output file
output_file = 'ROCKPROPS_UDF_Javelin_loader_March2023.XLSX'
output_dir = r'.\Outputs\\'

def create_output_loader(outfile, loading_template):
    if os.path.exists(outfile):
        print('Deleting existing Javelin ROCKPROPS loading workbook')
        os.remove(outfile)
    print('Creating an empty load template')
    shutil.copyfile(loading_template, outfile)
    return(outfile)


def add_borehole_interval_collection(inputs, workbook):
    # Inputs are: borehole_id, borehole_name, javelin[data]_location, instrument number, ground_dep_ref, acquisition_date
    ws = workbook['BOREHOLE INTERVAL COLLECTION']
    collection_name = inputs[1] + COLLECTION_NAME_POSTFIX
    new_row = [
        inputs[0],                              # Borehole_id aka borehole eno
        inputs[1],                              # Borehole_name
        inputs[4],                              # Ground_dep_ref
        '',                                     # Interval collection id
        collection_name,                        # Collection name
        COLLECTION_TYPE,                        # Collection type
        ORIGINATOR_NUMBER,                      # Originator number
        '',                                     # Preferred collection
        '',                                     # Collection source document id
        '',                                     # Source comments
    ]
    print(new_row)

    ws.append(new_row)
    return workbook, collection_name


def add_downhole_intervals(working_info, collection_name, javelin_results, workbook):
    intervals = workbook['DOWNHOLE INTERVALS']
    for idx, row in javelin_results.iterrows():
        interval_id = working_info[1] + '.jav.' + str(idx + 1)
        depth = row['depth']
        new_row = [
            '',                         # Collection no.
            collection_name,            # Collection name
            '',                         # Interval no.
            interval_id,                # Interval ID
            depth,                      # Interval start depth
            depth,                      # Interval end depth
            INTERVAL_UNIT_OF_MEASURE,   # Interval unit of measure
        ]
        intervals.append(new_row)
        workbook = add_samples(working_info, collection_name, interval_id, workbook, row)
    return workbook


def add_samples(working_info, collection_name, interval_id, workbook, jav_results_row):
    samples = workbook['SAMPLES']
    sample_id = interval_id + '.results'
    new_row = [
        working_info[0],          # Borehole eno
        collection_name,    # Collection name
        '',                 # Intervalno
        '',                 # Sampleno
        interval_id,        # Interval id
        sample_id,          # Sample id
        working_info[5],          # Acquisition date
        '',                 # Parent sampleno
        '',                 # Parent sample id
        ANO,                # ANO
        ACCESS_CODE,        # Access code
        '',                 # Confidential until date
        QA_STATUS,          # QA status
        ACTIVITY_CODE,      # Activity code
        SAMPLE_TYPE,        # Sample type
        SAMPLING_METHOD,    # Sampling method
        MATERIAL_CLASS,     # Material class
        '',                 # Procedure no
        '',                 # Project no
        '',                 # Refid
        '',                 # Other id
        '',                 # Specimen storage location
        '',                 # Storage date
        '',                 # ISGN
        '',                 # Specimen mass
        '',                 # Mass UOM
        '',                 # Source
    ]
    
    samples.append(new_row)
    workbook = add_scalar_properties(working_info, workbook, sample_id, jav_results_row)
    return workbook


def add_scalar_properties(working_info, workbook, sample_id, jav_results_row):
    scalar_properties = workbook['SCALAR PROPERTIES']
    jav_probe = working_info['instrument']
    process_lookup = {}
    if jav_probe == 350:
        process_lookup = PROCESS_LOOKUP_JP350
    elif jav_probe == 238:
        process_lookup = PROCESS_LOOKUP_JP238
    else:
        raise Exception('Unknown Javelin probe type. The value found was: {}'.format(jav_probe))
    
    for process in process_lookup:
        measurement = jav_results_row[process]
        if np.isnan(measurement):
            measurement = NULL_VALUE
        new_row = [
            sample_id,                  # SAMPLEID
            '',                         # Sample number
            '',                         # Resultno
            '',                         # Resultid
            ACCESS_CODE,                # Access code
            '',                         # Confidential until date
            QA_STATUS,                  # QA status
            ORIGNO,                     # Originator
            SOURCE_TYPE,                # Source type
            SOURCE,                     # Source
            LOAD_APPROVED,              # Load approved
            process_lookup[process][0], # Process type (PROCESS.PROCESSNO)
            process_lookup[process][1], # Petrophysical property
            measurement,                # Value
            process_lookup[process][2], # UOM
            process_lookup[process][3], # Result qualifier
            process_lookup[process][4], # uncertainty type
            '',                         # UNCERTAINTY VALUE
            '',                         # UNCERTAINTY UOM
            '',                         # RESULT DATE/TIME
            '',                         # OBSERVATION LOCATION
            process_lookup[process][5], # REMARKS
        ]
        
        print(new_row)
        scalar_properties.append(new_row)
        print(scalar_properties.max_row)

    return(workbook)


def main():
    # Read processing configuration information
    javelin_metadata = pd.read_excel(metadata_file)
    
    # Create output file
    output_path = output_dir + output_file
    working_loader = create_output_loader(output_path, template_file)
    wb = load_workbook(filename=working_loader)
    
    for row_nr in range(javelin_metadata.shape[0]):
        working_info = (javelin_metadata.iloc[row_nr])
        javelin_data = pd.read_csv(
            working_info['javelin_location'],
            delim_whitespace=True
        )

        #if row_nr == 0:
            # Create output file
            #bore_output_file = r'.\Outputs\\' + working_info[1] + '.jav.' + output_file
            #working_loader = create_output_loader(bore_output_file, template_file)
            #wb = load_workbook(filename=working_loader)
        

        print('\nPopulate BOREHOLE INTERVAL COLLECTION')
        wb, collection_name = add_borehole_interval_collection(working_info, wb)
        wb = add_downhole_intervals(working_info, collection_name, javelin_data, wb)

    wb.save(working_loader)

"""


if __name__ == "__main__":
    main()



